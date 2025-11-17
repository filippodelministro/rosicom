from openpyxl import load_workbook
import pandas as pd
import re

file_path = "source/ALL_ROSTERS.xlsx"

# Define team blocks (column, start_row, end_row)
team_blocks = [
    ('A', 5, 32),    # 1st team
    ('F', 5, 32),    # 2nd team
    ('A', 34, 61),   # 3rd team
    ('F', 34, 61),   # 4th team
    ('A', 63, 90),   # 5th team
    ('F', 63, 90),   # 6th team
    ('A', 92, 119),  # 7th team
    ('F', 92, 119),  # 8th team
    ('A', 121, 148), # 9th team
    ('F', 121, 148)  # 10th team
]

def get_all_teams(file_path):
    wb = load_workbook(filename=file_path, data_only=True)
    ws = wb.active

    team_names = []
    for col, start_row, end_row in team_blocks:
        team_name = ws[f'{col}{start_row}'].value
        team_names.append(team_name)
    return team_names

def get_team_data(file_path, team_index):
    wb = load_workbook(filename=file_path, data_only=True)
    ws = wb.active

    col, start_row, end_row = team_blocks[team_index]

    # Team name
    team_name = ws[f'{col}{start_row}'].value

    # Coins left
    coin_str = ws[f'{col}{end_row}'].value
    coins_left = int(re.search(r'\d+', coin_str).group())

    # Players
    df_player = pd.DataFrame(columns=['Team', 'Role', 'Player', 'PlayerTeam', 'Cost', 'CoinsLeft'])
    for row in range(start_row+2, end_row):
        role = ws[f'{col}{row}'].value
        player_name = ws[f'{col}{row}'].offset(column=1).value
        player_team = ws[f'{col}{row}'].offset(column=2).value
        cost = ws[f'{col}{row}'].offset(column=3).value
        df_player.loc[len(df_player)] = [team_name, role, player_name, player_team, cost, coins_left]

    return df_player

# --- Build one DataFrame for all teams ---
all_players_df = pd.DataFrame(columns=['Team', 'Role', 'Player', 'PlayerTeam', 'Cost', 'CoinsLeft'])

for i in range(len(team_blocks)):
    team_df = get_team_data(file_path, i)
    all_players_df = pd.concat([all_players_df, team_df], ignore_index=True)

# --- Example usage ---
print(all_players_df)
milan_players = all_players_df[all_players_df['PlayerTeam']=='Mil']
print("\nPlayers from Milan:", milan_players)

# print(all_players_df.head(10))
# print("\nShape of DataFrame:", all_players_df.shape)
# print("Columns:", all_players_df.columns)